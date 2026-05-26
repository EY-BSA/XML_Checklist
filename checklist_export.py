"""
checklist_export.py — 체크리스트 결과를 XLSX 템플릿으로 저장
app.py 의 api_export 로직을 독립 함수로 분리 (GUI / 웹 공용)
"""
from __future__ import annotations

import io
import re
import zipfile
from collections import OrderedDict
from pathlib import Path

_DRAWING_PAT = re.compile(rb'<(?:legacyDrawing|drawing)[^>]*/>')

# role_code 접두어 → 본문 재무제표 순서 (그 외는 주석으로 처리)
_STMT_ORDER = {'D2': 0, 'D4': 1, 'D6': 2, 'D5': 3}


def _issue_sort_key(iss):
    """연결→별도, 재무상태표→손익→자본변동→현금흐름→주석(번호순)"""
    rc = iss.role_code or ''

    # 연결(0) / 별도(1) / 미분류(2)
    if iss.is_consolidated is True:
        consol = 0
    elif iss.is_consolidated is False:
        consol = 1
    else:
        consol = 2

    # 본문 재무제표 순서 (D2/D4/D6/D5) vs 주석(4)
    stmt = _STMT_ORDER.get(rc[:2], 4)

    # 주석이면 role_name_ko 앞 숫자("24. 법인세비용" → 24) 추출
    if stmt == 4:
        m = re.match(r'(\d+)\.', (iss.role_name_ko or '').strip())
        note_num = int(m.group(1)) if m else 9999
    else:
        note_num = 0

    return (consol, stmt, note_num, rc)


def _role_def(iss) -> str:
    rc = iss.role_code; ko = iss.role_name_ko; en = iss.role_name_en
    if rc and ko:
        return f"[{rc}] {ko} | {en}" if en else f"[{rc}] {ko}"
    return ko or rc


def _tbl_name(iss) -> str:
    if iss.table_name_ko:
        return iss.table_name_ko
    rc = iss.role_code; ko = iss.role_name_ko
    if rc and ko:
        return f"[{rc}] {ko}"
    return ko or rc


def _write_std(ws, ri: int, iss) -> None:
    ws.cell(ri, 2, _role_def(iss))
    ws.cell(ri, 3, _tbl_name(iss))
    ws.cell(ri, 4, iss.prefix)
    ws.cell(ri, 5, iss.element_name)
    ws.cell(ri, 6, iss.label_ko)
    ws.cell(ri, 7, iss.label_en)
    ws.cell(ri, 8, iss.label_role)
    ws.cell(ri, 9, iss.data_type)
    ws.cell(ri, 10, iss.period)


def _find_sheet_xml(template_path: str, sheet_name: str) -> str:
    with zipfile.ZipFile(template_path, 'r') as z:
        wb_xml   = z.read('xl/workbook.xml').decode('utf-8')
        rels_xml = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    escaped = re.escape(sheet_name)
    m = re.search(rf'<sheet\b[^>]*\bname="{escaped}"[^>]*/>', wb_xml)
    if not m:
        raise KeyError(f'workbook.xml에서 시트 찾기 실패: {sheet_name}')
    rid_m = re.search(r'r:id="([^"]+)"', m.group(0))
    if not rid_m:
        raise KeyError(f'r:id 추출 실패: {sheet_name}')
    rid = rid_m.group(1)
    m2 = re.search(rf'<Relationship\b[^>]*\bId="{re.escape(rid)}"[^>]*/>', rels_xml)
    if not m2:
        raise KeyError(f'rId 찾기 실패: {rid}')
    target_m = re.search(r'Target="worksheets/([^"]+)"', m2.group(0))
    if not target_m:
        raise KeyError(f'Target 추출 실패: {rid}')
    return f'xl/worksheets/{target_m.group(1)}'


def export_checklist(results: OrderedDict, template_path: str, out_path: str) -> str:
    """
    체크리스트 결과를 XBRL_CoE_Checklist_Result 템플릿 형식으로 저장.

    Parameters
    ----------
    results       : run_all_checks() 반환 OrderedDict
    template_path : XBRL_CoE_Checklist_Result.xlsx 경로
    out_path      : 저장할 .xlsx 경로

    Returns str: 저장된 파일 경로
    """
    import openpyxl

    # ── Step 1: openpyxl 로 데이터 기록 ─────────────────────────────────
    wb = openpyxl.load_workbook(template_path)
    modified: set[str] = set()

    for chk_id, chk in results.items():
        sheet_name = chk.sheet[:31]
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        if ws.max_row >= 14:
            for row in ws.iter_rows(min_row=14, max_row=ws.max_row):
                for cell in row:
                    cell.value = None

        ws['B11'] = chk.issue_count
        modified.add(sheet_name)

        sorted_issues = sorted(chk.issues, key=_issue_sort_key)
        for ri, iss in enumerate(sorted_issues[:500], 14):
            if chk_id == '5-6':
                ws.cell(ri, 2, _role_def(iss))
                ws.cell(ri, 3, _tbl_name(iss))
                ws.cell(ri, 4, '단위미표시')
            elif chk_id == '7-1':
                _write_std(ws, ri, iss)
                ws.cell(ri, 11, iss.dart_negate)
                ws.cell(ri, 12, iss.client_negate)
            else:
                _write_std(ws, ri, iss)

    mod_buf = io.BytesIO()
    wb.save(mod_buf)
    mod_buf.seek(0)

    # ── Step 2: zipfile 로 템플릿 구조 보존 (drawing 등) ─────────────────
    sheet_xml_map: dict[str, str] = {}
    for sname in modified:
        try:
            sheet_xml_map[sname] = _find_sheet_xml(template_path, sname)
        except KeyError:
            pass

    preserve = {'xl/sharedStrings.xml', 'xl/styles.xml'}
    preserve.update(sheet_xml_map.values())
    for xml_path in sheet_xml_map.values():
        parts = xml_path.rsplit('/', 1)
        preserve.add(f'{parts[0]}/_rels/{parts[1]}.rels')

    drawing_tags_map: dict[str, list] = {}
    with zipfile.ZipFile(template_path, 'r') as tmpl_zip:
        for xml_path in sheet_xml_map.values():
            try:
                raw = tmpl_zip.read(xml_path)
                drawing_tags_map[xml_path] = _DRAWING_PAT.findall(raw)
            except Exception:
                drawing_tags_map[xml_path] = []

    out_buf = io.BytesIO()
    with zipfile.ZipFile(template_path, 'r') as tmpl_zip, \
         zipfile.ZipFile(mod_buf, 'r') as mod_zip:
        mod_names = set(mod_zip.namelist())
        with zipfile.ZipFile(out_buf, 'w', zipfile.ZIP_DEFLATED) as out_zip:
            for info in tmpl_zip.infolist():
                if info.filename == 'xl/calcChain.xml':
                    continue
                raw = tmpl_zip.read(info.filename)

                if info.filename in preserve and info.filename in mod_names:
                    raw = mod_zip.read(info.filename)
                    drawing_tags = drawing_tags_map.get(info.filename, [])
                    if drawing_tags and not _DRAWING_PAT.search(raw):
                        inject = b''.join(drawing_tags)
                        ws_tag_end = raw.find(b'>', raw.find(b'<worksheet'))
                        if b'xmlns:r=' not in raw[:ws_tag_end]:
                            ns = b' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
                            inject = re.sub(
                                rb'<((?:legacyDrawing|drawing)) ',
                                rb'<\1' + ns + rb' ',
                                inject,
                            )
                        raw = raw.replace(b'</worksheet>', inject + b'</worksheet>')

                if info.filename == '[Content_Types].xml':
                    raw = re.sub(rb'<Override[^>]*calcChain[^>]*/>', b'', raw)
                if info.filename == 'xl/workbook.xml':
                    raw = re.sub(
                        rb'<calcPr([^/]*)/>', rb'<calcPr\1 fullCalcOnLoad="1"/>', raw
                    )
                out_zip.writestr(info, raw)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    Path(out_path).write_bytes(out_buf.getvalue())
    return out_path
