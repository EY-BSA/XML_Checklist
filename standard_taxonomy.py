"""
standard_taxonomy.py
금감원 DART XBRL 표준 택사노미 파일 파서
Parses 'Presentation Link' and 'Concepts' sheets from dart_taxonomy.xlsm
"""
import re
from typing import Dict, Set, Optional

_CODE_RE     = re.compile(r'\[([A-Za-z]{1,3}X?\d{4,})\]')
_URI_CODE_RE = re.compile(r'[/_-]([A-Za-z]{1,3}X?\d{4,})$')

# Standard table code sets used by checks 4-1 ~ 4-6
STD_CF_TABLE_CODES:  Set[str] = {'D851100', 'D851105',
                                   'DX520000', 'DX520005',
                                   'D510000',  'D510005',
                                   'DI520000', 'DI520005',
                                   'D520000',  'D520005'}
STD_SGA_TABLE_CODES: Set[str] = {'D834310', 'D834315','DX830000'}
STD_RP_TABLE_CODES:  Set[str] = {'D818000', 'D818005', 'DX837000'}

_SKIP_MARKERS = {'LinkRole', 'Definition', 'prefix', 'None', ''}


class StandardTaxonomy:
    """Parsed lookup tables from the FSS standard taxonomy file."""

    def __init__(self):
        # element_name → set of table codes it appears in
        self.element_tables: Dict[str, Set[str]] = {}
        # element_name → 'negate' | '-'  (from preferredLabel in Presentation Link)
        self.dart_negate:    Dict[str, str]       = {}
        # element_name → {balance, periodType, type, abstract}
        self.element_props:  Dict[str, dict]      = {}
        # table_code → set of valid Axis/Domain/Member element names (from Axis_Domain_Check.xlsx)
        self.axis_members:   Dict[str, Set[str]]  = {}
        # 4-x 전용요소 집합: Axis_Domain_Check.xlsx TABLE_NUMBER 필터 기반 (R 인풋)
        self.cf_excl:  Set[str] = set()
        self.sga_excl: Set[str] = set()
        self.rp_excl:  Set[str] = set()

    # -- convenience sets (populated after load) --
    @property
    def cf_elements(self) -> Set[str]:
        return self._elem_set_for(STD_CF_TABLE_CODES)

    @property
    def sga_elements(self) -> Set[str]:
        return self._elem_set_for(STD_SGA_TABLE_CODES)

    @property
    def rp_elements(self) -> Set[str]:
        return self._elem_set_for(STD_RP_TABLE_CODES)

    def _elem_set_for(self, codes: Set[str]) -> Set[str]:
        return {e for e, tbl_set in self.element_tables.items()
                if tbl_set & codes}

    def get_dart_negate(self, element_name: str) -> str:
        """Return 'negate' if element has negatedLabel in standard taxonomy, else '-'."""
        return self.dart_negate.get(element_name, '-')

    # Axis_Domain_Check 기반 전용요소 조회 (4-x 체크용)
    def element_in_cf(self, element_name: str) -> bool:
        if self.cf_excl:
            return element_name in self.cf_excl
        tables = self.element_tables.get(element_name, set())
        return bool(tables & STD_CF_TABLE_CODES)

    def element_in_sga(self, element_name: str) -> bool:
        if self.sga_excl:
            return element_name in self.sga_excl
        tables = self.element_tables.get(element_name, set())
        return bool(tables & STD_SGA_TABLE_CODES)

    def element_in_rp(self, element_name: str) -> bool:
        if self.rp_excl:
            return element_name in self.rp_excl
        tables = self.element_tables.get(element_name, set())
        return bool(tables & STD_RP_TABLE_CODES)


def _parse_presentation_link(ws, std: StandardTaxonomy):
    current_code = ''
    for row in ws.iter_rows(values_only=True):
        marker = str(row[3]) if row[3] is not None else ''
        if marker == 'LinkRole':
            # Extract table code from URI in col[4]  e.g. '...role-D210000'
            uri = str(row[4]) if row[4] else ''
            m = _URI_CODE_RE.search(uri)
            current_code = m.group(1) if m else ''
            continue
        if marker in _SKIP_MARKERS:
            continue

        # Data row: row[3] = prefix, row[4] = element name, row[12] = preferredLabel
        elem = str(row[4]) if row[4] else ''
        if not elem or elem == 'None':
            continue

        if current_code:
            std.element_tables.setdefault(elem, set()).add(current_code)

        pref = str(row[12]) if row[12] else ''
        if 'negated' in pref.lower():
            # Only set to 'negate' — never downgrade once set
            std.dart_negate[elem] = 'negate'
        elif elem not in std.dart_negate:
            std.dart_negate[elem] = '-'


def _parse_concepts(ws, std: StandardTaxonomy):
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue  # skip header row
        name = str(row[2]) if row[2] else ''
        if not name or name == 'None':
            continue
        std.element_props[name] = {
            'prefix':     str(row[1]) if row[1] else '',
            'type':       str(row[4]) if row[4] else '',
            'balance':    str(row[7]).lower() if row[7] else '',
            'periodType': str(row[8]) if row[8] else '',
            'abstract':   bool(row[9]),
        }


def enrich_axis_domain_check(std: StandardTaxonomy, path: str) -> None:
    """Axis_Domain_Check.xlsx를 읽어 std.axis_members 와 4-x 전용요소 집합을 채운다.
    메인 포맷: result 시트, Table_Number / Name / Axis_Domain 컬럼.
    0511 포맷: 첫 번째 시트, Table_Number / KEY / Axis_Domain 컬럼
               (Name은 KEY 컬럼에서 'Axis_Name-Name' 형태로 추출).
    """
    import pandas as pd
    raw = pd.read_excel(path)
    raw.columns = raw.columns.astype(str).str.strip()

    if 'Name' in raw.columns:
        # 메인 포맷
        df = raw[['Table_Number', 'Name', 'Axis_Domain']].copy()
    elif 'KEY' in raw.columns:
        # 0511 포맷: KEY = "Axis_Name-Name", Name은 '-' 기준 두 번째 토큰
        df = raw[['Table_Number', 'KEY', 'Axis_Domain']].copy()
        df['Name'] = df['KEY'].astype(str).str.split('-', n=1).str[1].str.strip()
    else:
        return

    df['Table_Number'] = df['Table_Number'].astype(str).str.strip()
    df['Name']         = df['Name'].astype(str).str.strip()

    for _, row in df[df['Axis_Domain'].notna()].iterrows():
        table = row['Table_Number']
        name  = row['Name']
        if table and name:
            std.axis_members.setdefault(table, set()).add(name)

    std.cf_excl  = set(df[df['Table_Number'].isin(STD_CF_TABLE_CODES)]['Name'])
    std.sga_excl = set(df[df['Table_Number'].isin(STD_SGA_TABLE_CODES)]['Name'])
    std.rp_excl  = set(df[df['Table_Number'].isin(STD_RP_TABLE_CODES)]['Name'])


def enrich_dart_negate_check(std: StandardTaxonomy, path: str) -> None:
    """DART_Negate_Check.xlsx를 읽어 std.dart_negate 를 덮어쓴다.
    메인 포맷: result 시트, Name / Negate(negate|-) 컬럼.
    0511 포맷: 첫 번째 시트, Taxonomy ID(Prefix_Name) / DART_Negate 컬럼.
    """
    import pandas as pd
    raw = pd.read_excel(path)
    raw.columns = raw.columns.astype(str).str.strip()

    if 'Name' in raw.columns and 'Negate' in raw.columns:
        # 메인 포맷
        for _, row in raw.iterrows():
            name   = str(row['Name']).strip()
            negate = str(row['Negate']).strip()
            if name:
                std.dart_negate[name] = negate if negate == 'negate' else '-'
    elif 'Taxonomy ID' in raw.columns and 'DART_Negate' in raw.columns:
        # 0511 포맷: Taxonomy ID = "Prefix_Name" → Name은 '_' 기준 두 번째 토큰
        for _, row in raw.iterrows():
            tax_id = str(row['Taxonomy ID']).strip()
            negate = str(row['DART_Negate']).strip()
            if '_' in tax_id:
                name = tax_id.split('_', 1)[1].strip()
                if name:
                    std.dart_negate[name] = negate if negate == 'negate' else '-'


_singleton: Optional[StandardTaxonomy] = None
_singleton_path: str = ''


def get_standard_taxonomy(path: str) -> StandardTaxonomy:
    """Load and cache the standard taxonomy. Returns singleton after first call."""
    global _singleton, _singleton_path
    if _singleton is not None and _singleton_path == path:
        return _singleton

    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, keep_vba=False)

    std = StandardTaxonomy()
    _parse_presentation_link(wb['Presentation Link'], std)
    _parse_concepts(wb['Concepts'], std)
    wb.close()

    _singleton = std
    _singleton_path = path
    return std
