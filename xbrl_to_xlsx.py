"""
xbrl_to_xlsx.py
────────────────────────────────────────────────────────────────────────────
XBRL 원문 ZIP (.xsd + _pre.xml + _lab-ko.xml + _lab-en.xml + .xbrl)
   ↓
Presentation 시트 1장 xlsx 생성.

생성되는 시트:
  1. Presentation  — Presentation 트리 평탄화 결과
                     [role, table, id, label-ko, label-en, depth, order,
                      pref_label, DataType, Period, Decimal]
"""
from __future__ import annotations

import json
import re
import shutil
import tempfile
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

_TAXONOMY_JSON = Path(__file__).parent / "dart_taxonomy.json"
_taxonomy_cache: dict[str, dict[str, str]] | None = None


def _load_dart_taxonomy() -> dict[str, dict[str, str]]:
    global _taxonomy_cache
    if _taxonomy_cache is None:
        if _TAXONOMY_JSON.exists():
            _taxonomy_cache = json.loads(_TAXONOMY_JSON.read_text(encoding="utf-8"))
        else:
            _taxonomy_cache = {}
    return _taxonomy_cache

# ───────────────────────────────────────────────────────────────────────────
# Presentation 후처리
# ───────────────────────────────────────────────────────────────────────────
_LABEL_KO_EXCLUDE: frozenset[str] = frozenset({
    "연결 또는 별도 재무제표 [Table]",
    "연결 또는 별도 재무제표 [table]",
    "연결재무제표와 별도재무제표 [축]",
    "연결재무제표와 별도제무제표 [축]",
    "연결 또는 별도 재무제표 [domain]",
    "연결재무제표 [구성요소]",
    "별도재무제표 [구성요소]",
})


_ROLE_PREFIX_RE = re.compile(r"^\[[A-Z]\d+\]\s*")


def _role_to_table_name(role: str) -> str:
    """'[D210000] 재무상태표, 유동/비유동법 - 연결' → '재무상태표, 유동/비유동법 - 연결'"""
    return _ROLE_PREFIX_RE.sub("", role).strip()


def _postprocess_presentation(rows: list[dict]) -> list[dict]:
    # 1) id가 'Table'로 끝나는 행을 기준으로 테이블명 전파 (role 경계에서 초기화)
    current_role: str = ""
    current_table: str = ""
    for row in rows:
        if row["role"] != current_role:
            current_role = row["role"]
            current_table = ""
        if (row["id"] or "").endswith("Table"):
            current_table = row["label-ko"] or row["id"]
        row["table"] = current_table

    # 2) Table 행 바로 위 Abstract / TextBlock / Explanatory 행에도 동일 테이블명 소급
    for i, row in enumerate(rows):
        if (row["id"] or "").endswith("Table"):
            j = i - 1
            while j >= 0 and rows[j]["role"] == row["role"]:
                pid = rows[j]["id"] or ""
                if pid.endswith("Abstract") or pid.endswith("TextBlock") or pid.endswith("Explanatory"):
                    rows[j]["table"] = row["table"]
                    j -= 1
                else:
                    break

    # 3) table이 비어있는 행은 role에서 앞의 '[Dxxxxxx] ' 부분을 제거한 값으로 채움
    for row in rows:
        if not row["table"] and row["role"]:
            row["table"] = _role_to_table_name(row["role"])

    # 4) order 없는 행 삭제
    rows = [r for r in rows if r["order"] is not None]

    # 5) 특정 label-ko 행 삭제
    rows = [r for r in rows if r["label-ko"] not in _LABEL_KO_EXCLUDE]

    return rows


# ───────────────────────────────────────────────────────────────────────────
# 네임스페이스
# ───────────────────────────────────────────────────────────────────────────
NS = {
    "xsd":   "http://www.w3.org/2001/XMLSchema",
    "xbrli": "http://www.xbrl.org/2003/instance",
    "link":  "http://www.xbrl.org/2003/linkbase",
    "xlink": "http://www.w3.org/1999/xlink",
    "xml":   "http://www.w3.org/XML/1998/namespace",
}
XLINK_HREF  = f"{{{NS['xlink']}}}href"
XLINK_LABEL = f"{{{NS['xlink']}}}label"
XLINK_FROM  = f"{{{NS['xlink']}}}from"
XLINK_TO    = f"{{{NS['xlink']}}}to"
XLINK_ROLE  = f"{{{NS['xlink']}}}role"

DEFAULT_LABEL_ROLE = "http://www.xbrl.org/2003/role/label"

# ───────────────────────────────────────────────────────────────────────────
# 공통 유틸
# ───────────────────────────────────────────────────────────────────────────
def href_to_id(href: str) -> str:
    return href.split("#", 1)[1] if "#" in href else href



# ───────────────────────────────────────────────────────────────────────────
# 1) xsd 파싱 → Concepts 시트 데이터
# ───────────────────────────────────────────────────────────────────────────
def parse_xsd(path: str) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]], dict[str, str]]:
    """
    Returns
    -------
    concepts_rows : list of dict — Concepts 시트에 1:1
    concepts_by_id: dict[id → row] — 빠른 조회용
    role_def      : dict[roleURI → definition('[Dxxx] 한글 | English')]
    """
    tree = ET.parse(path)
    root = tree.getroot()

    rows: list[dict[str, Any]] = []
    by_id: dict[str, dict[str, Any]] = {}

    for el in root.iter(f"{{{NS['xsd']}}}element"):
        cid = el.get("id", "")
        if not cid:
            continue

        # substitutionGroup, type 는 'xbrli:xxx' / 'xbrldt:xxx' 형태 그대로 유지
        sub_group = el.get("substitutionGroup", "")
        ctype     = el.get("type", "")
        balance   = el.get(f"{{{NS['xbrli']}}}balance", "")
        period    = el.get(f"{{{NS['xbrli']}}}periodType", "")
        abstract  = el.get("abstract", "false")
        nillable  = el.get("nillable", "false")

        row = {
            "id":                cid,
            "name":              el.get("name", ""),
            "nillable":          nillable,
            "substitutionGroup": sub_group,
            "type":              ctype,
            "balance":           balance if balance else None,
            "periodType":        period if period else None,
            "abstract":          abstract,
        }
        rows.append(row)
        by_id[cid] = row

    role_def: dict[str, str] = {}
    for rt in root.iter(f"{{{NS['link']}}}roleType"):
        uri = rt.get("roleURI", "")
        defn = rt.find(f"{{{NS['link']}}}definition")
        if uri and defn is not None and defn.text:
            role_def[uri] = defn.text.strip()

    return rows, by_id, role_def


# ───────────────────────────────────────────────────────────────────────────
# 2) Label linkbase 파싱 → {concept_id: {role_uri: label_text}}
# ───────────────────────────────────────────────────────────────────────────
def parse_labels(path: str) -> dict[str, dict[str, str]]:
    tree = ET.parse(path)
    root = tree.getroot()

    loc_to_id: dict[str, str] = {}
    res_to_role_text: dict[str, tuple[str, str]] = {}
    arcs: list[tuple[str, str]] = []

    for ll in root.iter(f"{{{NS['link']}}}labelLink"):
        for loc in ll.findall(f"{{{NS['link']}}}loc"):
            href = loc.get(XLINK_HREF, "")
            lbl  = loc.get(XLINK_LABEL, "")
            cid  = href_to_id(href)
            if cid and lbl:
                loc_to_id[lbl] = cid
        for lab in ll.findall(f"{{{NS['link']}}}label"):
            lbl  = lab.get(XLINK_LABEL, "")
            role = lab.get(XLINK_ROLE, "")
            text = lab.text or ""
            if lbl:
                res_to_role_text[lbl] = (role, text)
        for arc in ll.findall(f"{{{NS['link']}}}labelArc"):
            f = arc.get(XLINK_FROM, "")
            t = arc.get(XLINK_TO, "")
            if f and t:
                arcs.append((f, t))

    labels: dict[str, dict[str, str]] = defaultdict(dict)
    for f, t in arcs:
        cid = loc_to_id.get(f)
        rt  = res_to_role_text.get(t)
        if cid and rt:
            role, text = rt
            labels[cid][role] = text
    return labels


def pick_label(labels_for_cid: dict[str, str], pref_role: str | None) -> str:
    """presentation row 의 label-ko / label-en 용 선택"""
    if not labels_for_cid:
        return ""
    if pref_role and pref_role in labels_for_cid:
        return labels_for_cid[pref_role]
    if DEFAULT_LABEL_ROLE in labels_for_cid:
        return labels_for_cid[DEFAULT_LABEL_ROLE]
    # fallback: 첫 번째 라벨
    return next(iter(labels_for_cid.values()))


# ───────────────────────────────────────────────────────────────────────────
# 3) Presentation linkbase → 평탄화된 row 리스트
# ───────────────────────────────────────────────────────────────────────────
def _role_label_only_ko(role_definition: str) -> str:
    """
    '[D210000] 재무상태표, 유동/비유동법 - 연결 | Statement ...'
    → '[D210000] 재무상태표, 유동/비유동법 - 연결'
    """
    if not role_definition:
        return ""
    return role_definition.split("|", 1)[0].strip()


def parse_presentation(
    path: str,
    labels_ko: dict[str, dict[str, str]],
    labels_en: dict[str, dict[str, str]],
    role_def_map: dict[str, str],
) -> list[dict[str, Any]]:
    tree = ET.parse(path)
    root = tree.getroot()
    rows: list[dict[str, Any]] = []

    for pl in root.iter(f"{{{NS['link']}}}presentationLink"):
        role_uri    = pl.get(XLINK_ROLE, "")
        role_def    = role_def_map.get(role_uri, role_uri)
        role_label  = _role_label_only_ko(role_def)

        loc_to_id: dict[str, str] = {}
        for loc in pl.findall(f"{{{NS['link']}}}loc"):
            loc_to_id[loc.get(XLINK_LABEL, "")] = href_to_id(loc.get(XLINK_HREF, ""))

        children: dict[str, list[tuple[str, float, str]]] = defaultdict(list)
        sources: set[str] = set()
        targets: set[str] = set()
        for arc in pl.findall(f"{{{NS['link']}}}presentationArc"):
            f = arc.get(XLINK_FROM, "")
            t = arc.get(XLINK_TO, "")
            try:
                order = float(arc.get("order", "0"))
            except ValueError:
                order = 0.0
            pref = arc.get("preferredLabel", "") or ""
            children[f].append((t, order, pref))
            sources.add(f); targets.add(t)
        for f in children:
            children[f].sort(key=lambda x: x[1])

        loc_order = {loc.get(XLINK_LABEL, ""): i
                     for i, loc in enumerate(pl.findall(f"{{{NS['link']}}}loc"))}
        roots = [lbl for lbl in loc_order
                 if lbl in sources and lbl not in targets]
        roots.sort(key=lambda x: loc_order.get(x, 0))

        def order_to_excel(o: float | None) -> Any:
            if o is None:
                return None
            if float(o).is_integer():
                return int(o)
            return o

        def emit(loc_label: str, depth: int, order_val: float | None, pref_role: str) -> None:
            cid = loc_to_id.get(loc_label, "")
            ko = pick_label(labels_ko.get(cid, {}), pref_role or None)
            en = pick_label(labels_en.get(cid, {}), pref_role or None)
            rows.append({
                "role":       role_label,
                "table":      "",
                "id":         cid,
                "label-ko":   ko,
                "label-en":   en,
                "depth":      depth,
                "order":      order_to_excel(order_val),
                "pref_label": pref_role or None,
            })

        def dfs(loc_label: str, depth: int, order_val: float | None, pref_role: str) -> None:
            emit(loc_label, depth, order_val, pref_role)
            for to_lbl, o, p in children.get(loc_label, []):
                dfs(to_lbl, depth + 1, o, p)

        for r in roots:
            dfs(r, 0, None, "")

    return rows


# ───────────────────────────────────────────────────────────────────────────
# 4) .xbrl 인스턴스 → fact 사전 (Decimal 채우기용)
# ───────────────────────────────────────────────────────────────────────────
def parse_instance_facts(xbrl_path: str) -> dict[str, list[dict[str, str]]]:
    """
    .xbrl 인스턴스에서 각 concept 의 fact 목록을 추출.

    Returns
    -------
    facts : dict[concept_name(=local name), list[{contextRef, decimals, value, unitRef}]]
    """
    facts: dict[str, list[dict[str, str]]] = defaultdict(list)
    if not xbrl_path or not Path(xbrl_path).exists():
        return facts

    NON_FACT_NS = {
        NS["xbrli"], NS["link"], NS["xlink"], NS["xsd"], NS["xml"],
        "http://xbrl.org/2006/xbrldi",
    }

    try:
        for _, el in ET.iterparse(xbrl_path, events=("end",)):
            tag = el.tag
            if not tag.startswith("{"):
                continue
            ns_uri, local = tag[1:].split("}", 1)
            if ns_uri in NON_FACT_NS:
                el.clear()
                continue
            ctx = el.get("contextRef")
            if not ctx:
                el.clear()
                continue
            facts[local].append({
                "contextRef": ctx,
                "decimals":   el.get("decimals", "") or "",
                "unitRef":    el.get("unitRef", "") or "",
                "value":      (el.text or "").strip(),
            })
            el.clear()
    except ET.ParseError:
        pass

    return facts


def pick_decimal(name: str, facts: dict[str, list[dict[str, str]]]) -> str:
    """
    당기(C로 시작) 컨텍스트 우선, 그 중 가장 작은(most rounded) decimals 반환.
    """
    fact_list = facts.get(name)
    if not fact_list:
        return ""
    candidates = [f for f in fact_list if f["contextRef"].startswith("C")] or fact_list
    numeric = []
    for f in candidates:
        try:
            numeric.append(int(f["decimals"]))
        except (ValueError, KeyError, TypeError):
            pass
    return str(min(numeric)) if numeric else ""


# ───────────────────────────────────────────────────────────────────────────
# 5) ZIP 안에서 XBRL 파일 자동 탐지 (pre를 기반으로)
# ───────────────────────────────────────────────────────────────────────────
def _resolve_xbrl_files(directory: str) -> dict[str, str]:
    result = {"xsd": "", "pre": "", "lab_ko": "", "lab_en": "", "xbrl": ""}
    for p in Path(directory).rglob("*"):
        if not p.is_file():
            continue
        name = p.name.lower()
        if name.endswith("_pre.xml"):
            result["pre"] = str(p)
        elif name.endswith("_lab-ko.xml") or name.endswith("_lab_ko.xml"):
            result["lab_ko"] = str(p)
        elif name.endswith("_lab-en.xml") or name.endswith("_lab_en.xml"):
            result["lab_en"] = str(p)
        elif name.endswith(".xbrl"):
            result["xbrl"] = str(p)
        elif name.endswith(".xsd") and not result["xsd"]:
            result["xsd"] = str(p)

    missing = [k for k in ("xsd", "pre", "lab_ko", "lab_en") if not result[k]]
    if missing:
        raise FileNotFoundError(
            "ZIP 안에서 다음 필수 파일을 찾지 못했습니다: " + ", ".join(missing)
        )
    return result


# ───────────────────────────────────────────────────────────────────────────
# 6) xlsx 작성
# ───────────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="FFD9E1F2")
BOLD = Font(bold=True)


def _split_concept_id(cid: str) -> tuple[str, str]:
    if not cid or "_" not in cid:
        return "", cid or ""
    pre, rest = cid.split("_", 1)
    return pre, rest



def _write_presentation(
    ws,
    presentation_rows: list[dict[str, Any]],
    concepts_by_id: dict[str, dict[str, Any]],
    facts: dict[str, list[dict[str, str]]],
) -> None:
    headers = ["role", "table", "id", "label-ko", "label-en", "depth", "order", "pref_label",
               "DataType", "Period", "Decimal"]
    ws.append(headers)
    for c in ws[1]:
        c.font = BOLD
        c.fill = HEADER_FILL

    for row in presentation_rows:
        cid = row["id"]
        info = concepts_by_id.get(cid, {})
        ctype = info.get("type", "") or ""
        data_type = ctype.split(":", 1)[-1] if ":" in ctype else ctype
        period = (info.get("periodType") or "").upper() or ""

        _, local_name = _split_concept_id(cid)
        decimal = pick_decimal(local_name, facts)

        ws.append([
            row["role"], row["table"], cid, row["label-ko"], row["label-en"],
            row["depth"], row["order"], row["pref_label"],
            data_type, period, decimal,
        ])

    widths = {"A": 50, "B": 50, "C": 60, "D": 50, "E": 50, "F": 8,
              "G": 8, "H": 50, "I": 22, "J": 12, "K": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


# ───────────────────────────────────────────────────────────────────────────
# 7) 메인 엔트리
# ───────────────────────────────────────────────────────────────────────────
def convert_zip_to_xlsx(zip_path: str, out_path: str) -> str:
    """
    XBRL ZIP 파일을 Presentation 1-시트 xlsx 로 변환.

    Returns: 실제로 저장된 xlsx 경로
    """
    tmp_dir = tempfile.mkdtemp(prefix="xbrl_")
    try:
        with zipfile.ZipFile(zip_path) as zf:
            zf.extractall(tmp_dir)

        files = _resolve_xbrl_files(tmp_dir)

        _, concepts_by_id, role_def = parse_xsd(files["xsd"])

        # dart_taxonomy.json 으로 ifrs-full / dart concept 보완
        for cid, info in _load_dart_taxonomy().items():
            if cid not in concepts_by_id:
                concepts_by_id[cid] = {
                    "id":                cid,
                    "name":              cid.split("_", 1)[-1] if "_" in cid else cid,
                    "type":              info.get("type", ""),
                    "periodType":        info.get("periodType", ""),
                    "balance":           info.get("balance") or None,
                    "abstract":          "false",
                    "nillable":          "false",
                    "substitutionGroup": "",
                }

        labels_ko = parse_labels(files["lab_ko"])
        labels_en = parse_labels(files["lab_en"])
        presentation_rows = _postprocess_presentation(
            parse_presentation(files["pre"], labels_ko, labels_en, role_def)
        )
        facts = parse_instance_facts(files["xbrl"]) if files["xbrl"] else {}

        wb = Workbook()
        wb.remove(wb.active)

        _write_presentation(
            wb.create_sheet("Presentation"),
            presentation_rows,
            concepts_by_id,
            facts,
        )

        Path(out_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return out_path

    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(
        description="XBRL ZIP → Presentation 1-시트 xlsx (DataType·Period·Decimal 포함)"
    )
    ap.add_argument("zip_path", help="입력 XBRL zip 파일")
    ap.add_argument("--out", help="출력 xlsx 경로 (기본: zip 과 동일 이름.xlsx)")
    args = ap.parse_args()

    in_path = Path(args.zip_path)
    out = args.out or str(in_path.with_suffix(".xlsx"))
    saved = convert_zip_to_xlsx(str(in_path), out)
    print(f"saved: {saved}")
